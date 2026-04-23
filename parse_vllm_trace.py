#!/usr/bin/env python3
"""Parse vLLM profiler traces to extract per-layer kernel breakdowns.

Important: vLLM traces must be captured with `VLLM_CUSTOM_SCOPES_FOR_PROFILING=1` to enable user annotation
           markers that provide scopes for gpu kernel events.

Usage:
    python parse_vllm_trace2.py <trace.json.gz>
        [--layer N]            Target transformer layer index (default: 8).
        [--percentile P]       The X-percentile batches to extract kernel stats from (default: 50).
        [--eager-trace <path>] Companion eager trace for extracting host module attributes.
                               For graph-captured batches like decode workloads in FULL_AND_PIECEWISE settings,
                               GPU kernels in a graph are packed into a single graph replay event.
                               An eager trace can be provided alongside with graphed traces to establish kernel
                               launches with host side launch modules from eager traces.
        [--output-prefix <str> Prefix for output XLSX filename.
"""

import argparse
import bisect
import gzip
import json
import os
import pickle
import re
import sys
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook


SPECIAL_KERNEL_LAUNCH_NAMES = {"hipmemcpyasync", "cudamemcpyasync"}

# vLLM iteration annotation pattern:
# execute_context_N(M)_generation_P(Q)
ITER_PATTERN = re.compile(r"execute_context_(\d+)\((\d+)\)_generation_(\d+)\((\d+)\)")

STAGE_PREFIX = "gpu_model_runner: "

NORM_NAMES = ("rmsnorm", "rms_norm", "layernorm", "layer_norm")

STABLE_KERNEL_PATTERNS = [
    "gemm",
    "gemv",
    "cublas",
    "cutlass",
    "triton_",
    "flash",
    "attention",
    "fused_moe",
    "grouped_gemm",
]


@dataclass(frozen=True)
class IterationInfo:
    """Parsed vLLM iteration annotation."""

    num_ctx_reqs: int
    num_ctx_tokens: int
    num_gen_reqs: int
    num_gen_tokens: int
    batch_type: str  # "pure_prefill", "pure_decode", "mixed", "empty"


def parse_vllm_iteration(name: str) -> Optional[IterationInfo]:
    m = ITER_PATTERN.search(name)
    if not m:
        return None
    ctx_reqs, ctx_tok, gen_reqs, gen_tok = (
        int(m.group(1)),
        int(m.group(2)),
        int(m.group(3)),
        int(m.group(4)),
    )
    if ctx_reqs == 0 and gen_reqs == 0:
        batch_type = "empty"
    elif ctx_reqs > 0 and gen_reqs == 0:
        batch_type = "pure_prefill"
    elif ctx_reqs == 0 and gen_reqs > 0:
        batch_type = "pure_decode"
    else:
        batch_type = "mixed"
    return IterationInfo(ctx_reqs, ctx_tok, gen_reqs, gen_tok, batch_type)

_SLIM_ARGS_KEYS = frozenset({"kernel", "correlation"})


def _slim_one(e: Dict) -> Optional[Dict]:
    """Slim raw event down to only the nodes needed."""
    if e.get("ph") != "X":
        return None
    slimmed: Dict[str, Any] = {
        "ph": "X",
        "cat": e.get("cat", ""),
        "name": e.get("name", ""),
        "ts": float(e.get("ts", 0)),
        "dur": float(e.get("dur", 0)),
        "tid": e.get("tid"),
        "pid": e.get("pid"),
    }
    raw_args = e.get("args")
    if raw_args:
        slim_args = {k: raw_args[k] for k in _SLIM_ARGS_KEYS if k in raw_args}
        if slim_args:
            slimmed["args"] = slim_args
    return slimmed


def load_trace_events(filepath: str) -> Tuple[List[Dict], float]:
    try:
        import ijson  # type: ignore[import]

        return _load_streaming(filepath, ijson)
    except ImportError:
        import warnings

        warnings.warn(
            "ijson not installed — falling back to json.load which loads the "
            "full trace into memory before slimming.  Install ijson for lower "
            "peak memory: pip install ijson",
            stacklevel=2,
        )
        return _load_full_and_slim(filepath)


def _load_streaming(filepath: str, ijson: Any) -> Tuple[List[Dict], float]:
    opener = gzip.open if filepath.endswith(".gz") else open
    kept: List[Dict] = []
    min_ts = float("inf")
    with opener(filepath, "rb") as f:
        for e in ijson.items(f, "traceEvents.item"):
            slimmed = _slim_one(e)
            if slimmed is not None:
                ts = slimmed["ts"]
                if ts < min_ts:
                    min_ts = ts
                kept.append(slimmed)
    return kept, min_ts if min_ts != float("inf") else 0.0


def _load_full_and_slim(filepath: str) -> Tuple[List[Dict], float]:
    opener = gzip.open if filepath.endswith(".gz") else open
    with opener(filepath, "rt", encoding="utf-8") as f:
        data = json.load(f)
    raw: List[Dict] = data.pop("traceEvents", []) if isinstance(data, dict) else data
    del data
    kept: List[Dict] = []
    min_ts = float("inf")
    for i in range(len(raw)):
        slimmed = _slim_one(raw[i])
        raw[i] = None  # type: ignore[assignment]
        if slimmed is not None:
            ts = slimmed["ts"]
            if ts < min_ts:
                min_ts = ts
            kept.append(slimmed)
    raw.clear()
    return kept, min_ts if min_ts != float("inf") else 0.0


def get_trace_start_ts(events: List[Dict]) -> float:
    min_ts = float("inf")
    for e in events:
        ts = e.get("ts")
        if ts is not None and ts < min_ts:
            min_ts = ts
    return min_ts if min_ts != float("inf") else 0.0


def format_perfetto_ts(ts_us: float, trace_start_ts: float) -> str:
    # Format absolute trace timestamp to Perfetto UI format
    rel_us = ts_us - trace_start_ts
    if rel_us < 0:
        rel_us = 0.0
    total_seconds = rel_us / 1_000_000.0
    hours = int(total_seconds // 3600)
    minutes = int((total_seconds % 3600) // 60)
    seconds = total_seconds % 60
    return f"{hours:02d}:{minutes:02d}:{seconds:012.9f}"


def is_kernel_launch(name: str) -> bool:
    n = name.lower()
    return ("launch" in n and "kernel" in n) or n in SPECIAL_KERNEL_LAUNCH_NAMES


def is_stable_anchor_kernel(name: str) -> bool:
    n = name.lower()
    return any(p in n for p in STABLE_KERNEL_PATTERNS)


def get_kernel_names_for_module(
    module_event: Dict,
    runtime_launches: List[Dict],
    runtime_launch_ts: List[float],
    kernel_by_corr: Dict,
) -> List[str]:
    m_start = module_event.get("ts", 0)
    m_end = m_start + module_event.get("dur", 0)
    left = bisect.bisect_left(runtime_launch_ts, m_start)
    right = bisect.bisect_right(runtime_launch_ts, m_end)
    names = []
    for launch in runtime_launches[left:right]:
        corr = (launch.get("args") or {}).get("correlation")
        if corr is not None:
            for k in kernel_by_corr.get(corr, []):
                kname = k.get("name", "")
                if kname:
                    names.append(kname)
    return names

_CXXFILT_PATH: Optional[str] = None
_DEMANGLE_CACHE: Dict[str, str] = {}


def _find_cxxfilt() -> Optional[str]:
    """Find llvm-cxxfilt or c++filt binary."""
    global _CXXFILT_PATH
    if _CXXFILT_PATH is not None:
        return _CXXFILT_PATH or None

    import shutil
    import subprocess

    for cmd in ["llvm-cxxfilt", "c++filt"]:
        path = shutil.which(cmd)
        if path:
            _CXXFILT_PATH = path
            return path

    known_paths = [
        "/opt/rocm/llvm/bin/llvm-cxxfilt",
        "/usr/bin/llvm-cxxfilt",
        "/usr/local/bin/llvm-cxxfilt",
    ]
    for p in known_paths:
        if os.path.isfile(p):
            _CXXFILT_PATH = p
            return p

    search_dirs = ["/root/.triton/llvm", "/opt/rocm"]
    for d in search_dirs:
        if not os.path.isdir(d):
            continue
        try:
            result = subprocess.run(
                [
                    "find",
                    d,
                    "-maxdepth",
                    "5",
                    "-name",
                    "llvm-cxxfilt",
                    "-type",
                    "f",
                    "-print",
                    "-quit",
                ],
                capture_output=True,
                text=True,
                timeout=5,
            )
            found = result.stdout.strip()
            if found:
                _CXXFILT_PATH = found
                return found
        except (subprocess.TimeoutExpired, OSError):
            continue

    _CXXFILT_PATH = ""
    return None


def _demangle_batch(names: list[str]) -> None:
    mangled = [n for n in names if n.startswith("_Z") and n not in _DEMANGLE_CACHE]
    if not mangled:
        return

    cxxfilt = _find_cxxfilt()
    if not cxxfilt:
        for n in mangled:
            _DEMANGLE_CACHE[n] = n
        return

    import subprocess

    try:
        result = subprocess.run(
            [cxxfilt],
            input="\n".join(mangled),
            capture_output=True,
            text=True,
            timeout=30,
        )
        demangled_list = result.stdout.strip().split("\n")
        for orig, dem in zip(mangled, demangled_list):
            _DEMANGLE_CACHE[orig] = dem.strip() if dem.strip() else orig
    except (subprocess.TimeoutExpired, OSError):
        for n in mangled:
            _DEMANGLE_CACHE[n] = n


def demangle_kernel_name(name: str) -> str:
    if name in _DEMANGLE_CACHE:
        return _DEMANGLE_CACHE[name]
    _DEMANGLE_CACHE[name] = name
    return name


class EventIndex:
    """Pre-indexed events for fast time-range queries."""

    def __init__(self, events: List[Dict]):
        self.duration_events = [e for e in events if e.get("ph") == "X"]
        self.duration_events.sort(key=lambda x: x["ts"])

        self._is_kernel_launch = [
            is_kernel_launch(e.get("name", "")) for e in self.duration_events
        ]
        self._kernel_prefix_sum = [0]
        for is_kl in self._is_kernel_launch:
            self._kernel_prefix_sum.append(
                self._kernel_prefix_sum[-1] + (1 if is_kl else 0)
            )

    def _bisect_range(self, start_ts: float, end_ts: float):
        """Return (left, right) indices for events starting in [start_ts, end_ts]."""
        left = bisect.bisect_left(self.duration_events, start_ts, key=lambda e: e["ts"])
        right = bisect.bisect_right(self.duration_events, end_ts, key=lambda e: e["ts"])
        return left, right

    def events_in_range(self, start_ts: float, end_ts: float) -> List[Dict]:
        """Get all duration events within [start_ts, end_ts]."""
        left, right = self._bisect_range(start_ts, end_ts)
        return [
            e
            for e in self.duration_events[left:right]
            if e["ts"] + e.get("dur", 0) <= end_ts
        ]

    def count_kernel_launches_in_range(self, start_ts: float, end_ts: float) -> int:
        left, right = self._bisect_range(start_ts, end_ts)
        count = 0
        for i in range(left, right):
            e = self.duration_events[i]
            if e["ts"] + e.get("dur", 0) <= end_ts and self._is_kernel_launch[i]:
                count += 1
        return count

    def get_direct_children(self, parent: Dict) -> List[Dict]:
        p_ts = parent["ts"]
        p_end = p_ts + parent.get("dur", 0)

        candidates = [e for e in self.events_in_range(p_ts, p_end) if e is not parent]
        if not candidates:
            return []

        candidates_sorted = sorted(candidates, key=lambda x: -x.get("dur", 0))

        direct = []
        for i, c in enumerate(candidates_sorted):
            c_ts, c_dur = c["ts"], c.get("dur", 0)
            c_end = c_ts + c_dur
            is_nested = False
            for j in range(i):
                o = candidates_sorted[j]
                o_ts = o["ts"]
                o_end = o_ts + o.get("dur", 0)
                if c_ts >= o_ts and c_end <= o_end:
                    is_nested = True
                    break
            if not is_nested:
                direct.append(c)

        return sorted(direct, key=lambda x: x["ts"])

    def count_kernel_launches(self, event: Dict) -> int:
        e_ts = event["ts"]
        e_end = e_ts + event.get("dur", 0)
        return self.count_kernel_launches_in_range(e_ts, e_end)

    def has_kernel_launch(self, event: Dict) -> bool:
        return self.count_kernel_launches(event) > 0


def kernel_signature(name: str) -> str:
    n = name
    n = re.sub(r"^void\s+", "", n)

    if not n.startswith("_Z"):
        n = re.sub(r"<[^<>]*>", "", n)
        n = re.sub(r"<[^<>]*>", "", n)
        n = re.sub(r"\([^)]*\)", "", n)
        n = re.sub(r"MT\d+x\d+x\d+", "MT_x_x_", n)
        n = re.sub(r"MI\d+x\d+x\d+", "MI_x_x_", n)
        n = re.sub(r"_\d+$", "_N", n)
        return n.strip()

    n = re.sub(r"ILi-?\d+E", "ILi_E", n)
    n = re.sub(r"E(\d+)E", r"E_E", n)
    n = re.sub(r"DF\d+\w", "DF_x", n)
    n = re.sub(r"MT\d+x\d+x\d+", "MT_x_x_", n)
    n = re.sub(r"MI\d+x\d+x\d+", "MI_x_x_", n)
    return n


def find_layer_boundaries(
    kernel_names: List[str],
    min_num_repeats: int = 4,
) -> Tuple[int, int, int]:
    """Find the repeating layer pattern in a kernel name sequence.

    Returns ``(loop_start, period, num_repeats)``.
    """
    sigs = [kernel_signature(n) for n in kernel_names]
    n = len(sigs)
    if n < min_num_repeats:
        return (0, n, 1)

    best_period = n
    best_start = 0
    best_repeats = 1

    max_period = n // min_num_repeats
    for period in range(3, min(max_period + 1, 200)):
        for start in range(0, min(period + 1, n - period * min_num_repeats + 1)):
            pattern = sigs[start : start + period]
            repeats = 1
            pos = start + period
            while pos + period <= n:
                if sigs[pos : pos + period] == pattern:
                    repeats += 1
                    pos += period
                else:
                    break
            if repeats >= min_num_repeats and repeats > best_repeats:
                best_period = period
                best_start = start
                best_repeats = repeats

    if best_repeats > 1:
        for offset in range(best_period):
            idx = best_start + offset
            if idx < n:
                name_lower = kernel_names[idx].lower()
                if any(kw in name_lower for kw in NORM_NAMES):
                    if offset > 0:
                        best_start += offset
                        max_repeats = (n - best_start) // best_period
                        if best_repeats > max_repeats:
                            best_repeats = max_repeats
                    break

    if best_repeats > 1 and best_start >= best_period:
        pattern = sigs[best_start : best_start + best_period]
        while best_start >= best_period:
            if sigs[best_start - best_period : best_start] == pattern:
                best_start -= best_period
                best_repeats += 1
            else:
                break

    return (best_start, best_period, best_repeats)


def write_kernel_rows_to_sheet(
    ws,
    rows: List[List[Any]],
    section_label: str = "",
) -> None:
    if section_label:
        ws.append([section_label, "", "", ""])

    total_duration = sum(float(r[2]) for r in rows) if rows else 0.0
    for cpu_mod, kernel, dur in rows:
        dur_f = float(dur)
        pct = (dur_f / total_duration * 100) if total_duration > 0 else 0
        ws.append(
            [
                cpu_mod,
                demangle_kernel_name(kernel),
                round(dur_f, 3),
                round(pct, 1),
            ]
        )

    ws.append(["TOTAL", "", round(total_duration, 3), 100.0])
    ws.append([])


def write_consolidated_xlsx(
    output_xlsx: str,
    sections: Dict[str, List[List[Any]]],
) -> None:
    _demangle_batch(
        r[1] for groups in sections.values() for _label, rows in groups for r in rows
    )

    wb = Workbook()
    first_sheet = True
    kernel_stats: Dict[str, Tuple[int, float]] = {}
    grand_total: float = 0.0

    for sheet_name, groups in sections.items():
        if first_sheet:
            ws = wb.active
            ws.title = sheet_name
            first_sheet = False
        else:
            ws = wb.create_sheet(sheet_name)

        ws.append(["cpu_module", "gpu_kernel", "duration_us", "pct%"])

        for label, rows in groups:
            write_kernel_rows_to_sheet(ws, rows, section_label=label)
            for _, kernel, dur in rows:
                short_name = demangle_kernel_name(kernel)
                dur_f = float(dur)
                prev = kernel_stats.get(short_name, (0, 0.0))
                kernel_stats[short_name] = (prev[0] + 1, prev[1] + dur_f)
                grand_total += dur_f

    if kernel_stats:
        _write_kernel_summary_sheet(wb, kernel_stats, grand_total)

    wb.save(output_xlsx)
    sheet_names = list(sections.keys())
    print(f"XLSX written to: {output_xlsx} (sheets: {', '.join(sheet_names)})")


def _write_kernel_summary_sheet(
    wb: "Workbook",
    kernel_stats: Dict[str, Tuple[int, float]],
    total_duration: float,
) -> None:
    ws = wb.create_sheet("kernel_summary")
    ws.append(["gpu_kernel", "calls", "total_duration_us", "avg_duration_us", "pct%"])

    sorted_kernels = sorted(kernel_stats.items(), key=lambda x: x[1][1], reverse=True)
    for kernel_name, (count, total_dur) in sorted_kernels:
        avg_dur = total_dur / count
        pct = (total_dur / total_duration * 100) if total_duration > 0 else 0
        ws.append(
            [
                kernel_name,
                count,
                round(total_dur, 3),
                round(avg_dur, 3),
                round(pct, 1),
            ]
        )


@dataclass
class NormData:
    norm_gpu_indices: Optional[List[int]]
    norm_period: int
    norm_loop_start: int
    use_end_as_boundary: bool


@dataclass
class EagerForwardResult:
    gpu_kernels: List[Dict]
    module_map: Dict[int, str]
    norm_data: NormData
    cpu_fwd: Optional[Dict]


def detect_pids(events: List[Dict]) -> Tuple[Any, Any]:
    cpu_pid = next(
        (e["pid"] for e in events if e.get("cat") == "user_annotation"), None
    )
    gpu_pid = next((e["pid"] for e in events if e.get("cat") == "kernel"), None)
    return cpu_pid, gpu_pid


def collect_iteration_events(
    events: List[Dict], cpu_pid: Any
) -> Tuple[List[Dict], List[Dict]]:
    prefill, decode = [], []
    for e in events:
        if e.get("ph") != "X":
            continue
        if e.get("cat") != "user_annotation":
            continue
        if e.get("pid") != cpu_pid:
            continue
        info = parse_vllm_iteration(e.get("name", ""))
        if info is None:
            continue
        if info.batch_type == "pure_decode":
            decode.append(e)
        elif info.batch_type in ("pure_prefill", "mixed"):
            prefill.append(e)
    prefill.sort(key=lambda x: x["ts"])
    decode.sort(key=lambda x: x["ts"])
    return prefill, decode


def _find_cpu_fwd(events: List[Dict], iteration: Dict, cpu_pid: Any) -> Optional[Dict]:
    """Find user_annotation 'gpu_model_runner: forward' inside an iteration."""
    it_ts = iteration["ts"]
    it_end = it_ts + iteration.get("dur", 0)
    it_tid = iteration.get("tid")
    target = STAGE_PREFIX + "forward"
    for e in events:
        if e.get("ph") != "X":
            continue
        if e.get("cat") != "user_annotation":
            continue
        if e.get("name") != target:
            continue
        if e.get("pid") != cpu_pid:
            continue
        if e.get("tid") != it_tid:
            continue
        e_ts = e.get("ts", 0)
        if it_ts <= e_ts and e_ts + e.get("dur", 0) <= it_end:
            return e
    return None


def select_iteration(
    iters: List[Dict],
    events: List[Dict],
    cpu_pid: Any,
    percentile: float,
    label: str,
) -> Optional[Dict]:
    """Pick the iteration at the given percentile of forward duration."""
    if not iters:
        return None

    durations = []
    for it in iters:
        fwd = _find_cpu_fwd(events, it, cpu_pid)
        dur = fwd["dur"] if fwd else it.get("dur", 0)
        durations.append(dur)

    n = len(iters)
    sorted_indices = sorted(range(n), key=lambda i: durations[i])
    rank = int(round(percentile / 100.0 * (n - 1)))
    rank = max(0, min(rank, n - 1))
    sel_idx = sorted_indices[rank]

    p10 = durations[sorted_indices[max(0, int(0.10 * (n - 1)))]]
    p50 = durations[sorted_indices[max(0, int(0.50 * (n - 1)))]]
    p90 = durations[sorted_indices[max(0, int(0.90 * (n - 1)))]]
    print(
        f"  {label}: {n} iters, p10={p10:.0f}us p50={p50:.0f}us p90={p90:.0f}us | "
        f"selected p{percentile:.0f} => iter[{sel_idx}] dur={durations[sel_idx]:.0f}us"
    )
    return iters[sel_idx]


def _build_corr_to_gpu(events: List[Dict], gpu_pid: Any) -> Dict[int, Dict]:
    """Map correlation ID to GPU kernel event."""
    corr_to_gpu: Dict[int, Dict] = {}
    for e in events:
        if e.get("ph") != "X" or e.get("cat") != "kernel":
            continue
        if e.get("pid") != gpu_pid:
            continue
        corr = (e.get("args") or {}).get("correlation")
        if corr is not None:
            corr_to_gpu[corr] = e
    return corr_to_gpu


def _get_gpu_fwd_annotation(
    events: List[Dict], gpu_pid: Any, launch_ts: float
) -> Optional[Dict]:
    """Find gpu_user_annotation forward that starts at/after launch_ts."""
    target = STAGE_PREFIX + "forward"
    best: Optional[Dict] = None
    best_ts = float("inf")
    for e in events:
        if e.get("ph") != "X" or e.get("cat") != "gpu_user_annotation":
            continue
        if e.get("pid") != gpu_pid:
            continue
        if e.get("name") != target:
            continue
        e_ts = e.get("ts", 0)
        if e_ts >= launch_ts and e_ts < best_ts:
            best_ts = e_ts
            best = e
    return best


def get_forward_kernels(
    events: List[Dict],
    iteration: Dict,
    cpu_pid: Any,
    gpu_pid: Any,
    corr_to_gpu: Optional[Dict[int, Dict]] = None,
) -> Tuple[List[Dict], bool, Optional[Dict]]:
    """Extract GPU kernels for the forward pass. Returns (kernels, is_cudagraph, cpu_fwd)."""
    it_ts = iteration["ts"]
    it_end = it_ts + iteration.get("dur", 0)
    it_tid = iteration.get("tid")
    it_pid = iteration.get("pid")

    is_cudagraph = False
    graph_launch_ts: Optional[float] = None
    for e in events:
        if e.get("ph") != "X" or e.get("cat") != "cuda_runtime":
            continue
        if e.get("pid") != it_pid or e.get("tid") != it_tid:
            continue
        e_ts = e.get("ts", 0)
        if not (it_ts <= e_ts <= it_end):
            continue
        if "graphlaunch" in e.get("name", "").lower():
            is_cudagraph = True
            graph_launch_ts = e_ts
            break

    if is_cudagraph:
        gpu_fwd = _get_gpu_fwd_annotation(events, gpu_pid, graph_launch_ts or it_ts)
        if gpu_fwd is None:
            print(
                "  CUDAGraph: No gpu_user_annotation forward found, fall back to using host iteration range."
                "             The iteration range can lead to incorrect kernels captured. Please set"
                "             VLLM_CUSTOM_SCOPES_FOR_PROFILING=1 when capturing traces to provide gpu markers"
            )
            g_ts, g_end = it_ts, it_end
        else:
            g_ts = gpu_fwd["ts"]
            g_end = g_ts + gpu_fwd.get("dur", 0)
        kernels = [
            e
            for e in events
            if e.get("ph") == "X"
            and e.get("cat") == "kernel"
            and e.get("pid") == gpu_pid
            and g_ts <= e.get("ts", 0) <= g_end
        ]
        kernels.sort(key=lambda x: x["ts"])
        print(
            f"  CUDAGraph detected — gpu_fwd window [{g_ts:.0f}, {g_end:.0f}]: "
            f"{len(kernels)} kernels"
        )
        return kernels, True, None

    # Eager mode
    cpu_fwd = _find_cpu_fwd(events, iteration, cpu_pid)
    if cpu_fwd is None:
        print(
            "  No cpu forward stage event found."
            "  Please set VLLM_CUSTOM_SCOPES_FOR_PROFILING=1 when collecting traces."
        )
        return [], False, None

    fwd_ts = cpu_fwd["ts"]
    fwd_end = fwd_ts + cpu_fwd.get("dur", 0)
    fwd_tid = cpu_fwd.get("tid")
    fwd_pid = cpu_fwd.get("pid")

    rt_launches = [
        e
        for e in events
        if e.get("ph") == "X"
        and e.get("cat") == "cuda_runtime"
        and e.get("pid") == fwd_pid
        and e.get("tid") == fwd_tid
        and fwd_ts <= e.get("ts", 0) <= fwd_end
        and is_kernel_launch(e.get("name", ""))
    ]
    rt_launches.sort(key=lambda x: x["ts"])

    if not rt_launches:
        return [], False, cpu_fwd

    if corr_to_gpu is None:
        corr_to_gpu = _build_corr_to_gpu(events, gpu_pid)

    kernels: List[Dict] = []
    for rt in rt_launches:
        corr = (rt.get("args") or {}).get("correlation")
        if corr is not None:
            gpu_ev = corr_to_gpu.get(corr)
            if gpu_ev is not None:
                kernels.append(gpu_ev)

    return kernels, False, cpu_fwd


def build_module_map(
    events: List[Dict],
    cpu_fwd: Dict,
    gpu_kernels: List[Dict],
) -> Dict[int, str]:
    """Map GPU kernel index to cpu_module name."""
    fwd_ts = cpu_fwd["ts"]
    fwd_end = fwd_ts + cpu_fwd.get("dur", 0)
    fwd_tid = cpu_fwd.get("tid")
    fwd_pid = cpu_fwd.get("pid")

    cpu_ops = [
        e
        for e in events
        if e.get("ph") == "X"
        and e.get("cat") == "cpu_op"
        and e.get("pid") == fwd_pid
        and e.get("tid") == fwd_tid
        and fwd_ts <= e.get("ts", 0)
        and e.get("ts", 0) + e.get("dur", 0) <= fwd_end
    ]
    cpu_ops.sort(key=lambda e: (e["ts"], e.get("dur", 0)))

    rt_launches = [
        e
        for e in events
        if e.get("ph") == "X"
        and e.get("cat") == "cuda_runtime"
        and e.get("pid") == fwd_pid
        and e.get("tid") == fwd_tid
        and fwd_ts <= e.get("ts", 0) <= fwd_end
        and is_kernel_launch(e.get("name", ""))
    ]
    rt_launches.sort(key=lambda x: x["ts"])

    corr_to_rt: Dict[int, Dict] = {}
    for rt in rt_launches:
        corr = (rt.get("args") or {}).get("correlation")
        if corr is not None:
            corr_to_rt[corr] = rt

    def _narrowest_cpu_op(rt_event: Dict) -> str:
        rt_ts = rt_event["ts"]
        rt_end = rt_ts + rt_event.get("dur", 0)
        containing = [
            op
            for op in cpu_ops
            if op["ts"] <= rt_ts and rt_end <= op["ts"] + op.get("dur", 0)
        ]
        containing.sort(key=lambda op: op.get("dur", 0))
        if not containing:
            return ""
        for op in containing:
            name = op.get("name", "")
            if not name.startswith("aiter::"):
                return name
        return containing[0].get("name", "")

    index_to_mod: Dict[int, str] = {}
    for i, gk in enumerate(gpu_kernels):
        corr = (gk.get("args") or {}).get("correlation")
        if corr is None:
            continue
        rt = corr_to_rt.get(corr)
        if rt is None:
            continue
        name = _narrowest_cpu_op(rt)
        if name:
            index_to_mod[i] = name

    return index_to_mod


def _collect_norm_cpu_ops(events: List[Dict], cpu_fwd: Dict) -> List[Dict]:
    """Collect outermost norm cpu_op events via universal nesting filter."""
    fwd_ts = cpu_fwd["ts"]
    fwd_end = fwd_ts + cpu_fwd.get("dur", 0)
    fwd_tid = cpu_fwd.get("tid")
    fwd_pid = cpu_fwd.get("pid")

    candidates = [
        e
        for e in events
        if e.get("ph") == "X"
        and e.get("cat") == "cpu_op"
        and e.get("pid") == fwd_pid
        and e.get("tid") == fwd_tid
        and fwd_ts <= e.get("ts", 0)
        and e.get("ts", 0) + e.get("dur", 0) <= fwd_end
        and any(kw in e.get("name", "").lower() for kw in NORM_NAMES)
    ]
    candidates.sort(key=lambda x: x["ts"])

    intervals = [(e.get("ts", 0), e.get("ts", 0) + e.get("dur", 0)) for e in candidates]
    norms = []
    for i, ev in enumerate(candidates):
        ev_ts, ev_end = intervals[i]
        nested = any(
            j != i and intervals[j][0] <= ev_ts and ev_end <= intervals[j][1]
            for j in range(len(candidates))
        )
        if not nested:
            norms.append(ev)
    return norms


# Keywords that identify attention-type kernels for phase detection.
_PHASE_ATTN_KWS = (
    "attention",
    "flash",
    "fmha",
    "mla",
    "rope_concat",
    "rope",
    "kv_cache",
    "fused_qk",
    "kn_entry",
    "mla_dec",
    "mqa_logits",
)


def _detect_phase(
    gpu_kernels: List[Dict],
    positions: List[int],
    loop_start: int,
) -> bool:
    """Return True if attention appears before any standalone norm in the inter-norm gap.

    In a pre-norm transformer the correct-phase gap (input_layernorm → input_layernorm)
    has attention first; the wrong-phase gap crosses a layer boundary and contains a
    standalone input_layernorm before the next layer's attention.
    """
    _NORM_KWS = ("rmsnorm", "rms_norm", "layernorm", "layer_norm")
    in_region = [(i, p) for i, p in enumerate(positions) if p >= loop_start]
    if len(in_region) < 2:
        return True
    _, p_start = in_region[0]
    _, p_end = in_region[1]
    first_attn = first_standalone_norm = float("inf")
    for j, k in enumerate(gpu_kernels[p_start + 1 : p_end]):
        name = k.get("name", "").lower()
        is_attn = any(w in name for w in _PHASE_ATTN_KWS)
        is_norm = any(kw in name for kw in _NORM_KWS)
        if first_attn == float("inf") and is_attn:
            first_attn = j
        if first_standalone_norm == float("inf") and is_norm and not is_attn:
            first_standalone_norm = j
    if first_attn == float("inf") and first_standalone_norm == float("inf"):
        return True
    return first_attn < first_standalone_norm


def _scan_for_norm_boundaries(
    gpu_kernels: List[Dict],
    loop_start: int,
    detected_period: int,
    expected_layers: int,
) -> Optional["NormData"]:
    """Scan GPU kernels for norm-type kernels and identify layer boundaries."""
    slack = max(2, expected_layers // 8)

    sig_positions: Dict[str, List[int]] = {}
    for i, k in enumerate(gpu_kernels):
        name = k.get("name", "").lower()
        if not any(kw in name for kw in NORM_NAMES):
            continue
        sig = kernel_signature(k.get("name", ""))
        sig_positions.setdefault(sig, []).append(i)

    if not sig_positions:
        return None

    candidates: List[Tuple[List[int], Optional[List[int]], float]] = []
    for sig, positions in sig_positions.items():
        cnt = len(positions)
        first_offset = min(
            (
                p - loop_start
                for p in positions
                if loop_start <= p < loop_start + detected_period
            ),
            default=float("inf"),
        )
        if abs(cnt - expected_layers) <= slack:
            candidates.append((positions, None, first_offset))
        elif abs(cnt - 2 * expected_layers) <= 2 * slack:
            primary = positions[::2]
            alt = positions[1::2]
            if abs(len(primary) - expected_layers) <= slack:
                candidates.append((primary, alt, first_offset))

    if not candidates:
        return None

    candidates.sort(key=lambda x: x[2])
    best_positions, best_alt, best_offset = candidates[0]

    if not _detect_phase(gpu_kernels, best_positions, loop_start):
        if best_alt is not None and abs(len(best_alt) - expected_layers) <= slack:
            print("  Phase corrected: switched to odd-indexed norms (input_layernorm)")
            best_positions = best_alt
        else:
            print(
                "  Warning: phase appears wrong but cannot correct (no alternate indexing)"
            )

    return NormData(best_positions, 1, 0, False)


def find_layer_boundaries_v2(
    events: List[Dict],
    cpu_fwd: Optional[Dict],
    gpu_kernels: List[Dict],
    gpu_pid: Any = None,
    corr_to_gpu: Optional[Dict[int, Dict]] = None,
) -> "NormData":
    """Detect layer boundaries. Returns NormData with norm indices and period.

    Pass 1: CPU norm hierarchy (requires cpu_fwd — eager mode).
    Pass 2: GPU kernel signature fallback (CUDAGraph without companion).
    """
    if cpu_fwd is None:
        return _gpu_signature_fallback(gpu_kernels)

    norm_events = _collect_norm_cpu_ops(events, cpu_fwd)
    if len(norm_events) < 2:
        print("  Too few norm cpu_op events, falling back to GPU signature detection")
        return _gpu_signature_fallback(gpu_kernels)

    fwd_ts = cpu_fwd["ts"]
    fwd_end = fwd_ts + cpu_fwd.get("dur", 0)
    fwd_tid = cpu_fwd.get("tid")
    fwd_pid = cpu_fwd.get("pid")

    rt_launches = [
        e
        for e in events
        if e.get("ph") == "X"
        and e.get("cat") == "cuda_runtime"
        and e.get("pid") == fwd_pid
        and e.get("tid") == fwd_tid
        and fwd_ts <= e.get("ts", 0) <= fwd_end
        and is_kernel_launch(e.get("name", ""))
    ]
    rt_launches.sort(key=lambda x: x["ts"])
    rt_ts_list = [e.get("ts", 0) for e in rt_launches]

    corr_to_gpu_idx: Dict[int, int] = {}
    for gi, gk in enumerate(gpu_kernels):
        corr = (gk.get("args") or {}).get("correlation")
        if corr is not None:
            corr_to_gpu_idx[corr] = gi

    used_gpu: set = set()
    norm_gpu_indices: List[int] = []
    mapped_cpu_names: List[str] = []

    for norm_ev in norm_events:
        n_ts = norm_ev.get("ts", 0)
        n_end = n_ts + norm_ev.get("dur", 0)
        left = bisect.bisect_left(rt_ts_list, n_ts)
        right = bisect.bisect_right(rt_ts_list, n_end)

        candidates_gi: List[int] = []
        for rt in rt_launches[left:right]:
            corr = (rt.get("args") or {}).get("correlation")
            if corr is not None and corr in corr_to_gpu_idx:
                gi = corr_to_gpu_idx[corr]
                if gi not in used_gpu:
                    candidates_gi.append(gi)

        if not candidates_gi:
            continue

        norm_candidates = [
            gi
            for gi in candidates_gi
            if any(kw in gpu_kernels[gi].get("name", "").lower() for kw in NORM_NAMES)
        ]
        found_gi = norm_candidates[-1] if norm_candidates else candidates_gi[-1]
        norm_gpu_indices.append(found_gi)
        mapped_cpu_names.append(norm_ev.get("name", ""))
        used_gpu.add(found_gi)

    if len(norm_gpu_indices) < 2:
        print("  Too few norm GPU mappings, falling back to GPU signature detection")
        return _gpu_signature_fallback(gpu_kernels)

    norm_period = 0
    norm_loop_start = 0
    dominant_name = ""

    if len(mapped_cpu_names) >= 4:
        name_counts = Counter(mapped_cpu_names)
        dominant_name = name_counts.most_common(1)[0][0]

        min_loop_count = 3
        loop_names = [n for n in mapped_cpu_names if name_counts[n] > min_loop_count]
        if len(loop_names) < 4:
            loop_names = list(mapped_cpu_names)

        if len(set(loop_names)) <= 1:
            norm_period = 2
            for i, n in enumerate(mapped_cpu_names):
                if n == dominant_name:
                    norm_loop_start = i
                    break
        else:
            _ls, norm_period, _ = find_layer_boundaries(loop_names, min_num_repeats=4)
            filtered_indices = [
                i
                for i, n in enumerate(mapped_cpu_names)
                if name_counts[n] > min_loop_count
            ]
            norm_loop_start = (
                filtered_indices[_ls] if _ls < len(filtered_indices) else 0
            )

    use_end = "fused_allreduce" in dominant_name

    if not use_end and norm_period > 1 and norm_loop_start < len(mapped_cpu_names):
        period_slice = mapped_cpu_names[norm_loop_start : norm_loop_start + norm_period]
        _FUSED_KWS = ("fwd_with_add", "fused_allreduce", "add_rmsnorm", "add_rms")
        first_simple_pos = None
        for pi, pname in enumerate(period_slice):
            is_norm = any(kw in pname.lower() for kw in NORM_NAMES)
            is_fused = any(kw in pname.lower() for kw in _FUSED_KWS)
            if is_norm and not is_fused:
                first_simple_pos = pi
                break
        if first_simple_pos is not None and first_simple_pos > 1:
            norm_loop_start += first_simple_pos - 1

    num_layers = (
        (len(norm_gpu_indices) - norm_loop_start) // norm_period
        if norm_period > 0
        else 0
    )

    # Phase correction for CPU norm path
    if (
        not use_end
        and norm_period == 2
        and len(norm_gpu_indices) >= norm_loop_start + 2
    ):
        current_positions = norm_gpu_indices[norm_loop_start::norm_period]
        gpu_loop_start = current_positions[0] if current_positions else 0
        if not _detect_phase(gpu_kernels, current_positions, gpu_loop_start):
            alt_ls = norm_loop_start + 1
            if alt_ls < len(norm_gpu_indices):
                alt_num_layers = len(norm_gpu_indices[alt_ls::norm_period])
                if abs(alt_num_layers - num_layers) <= 1:
                    print(
                        "  Phase corrected: CPU norm boundaries shifted to alternate loop_start"
                    )
                    norm_loop_start = alt_ls
                    num_layers = alt_num_layers

    return NormData(norm_gpu_indices, norm_period, norm_loop_start, use_end)


def _gpu_signature_fallback(gpu_kernels: List[Dict]) -> "NormData":
    """Fallback: detect layer boundaries from GPU kernel name signatures."""
    kernel_names = [k.get("name", "") for k in gpu_kernels]
    loop_start, period, num_repeats = find_layer_boundaries(
        kernel_names, min_num_repeats=4
    )
    print(
        f"  GPU signature fallback: {len(kernel_names)} kernels, "
        f"period={period}, repeats={num_repeats}, loop_start={loop_start}"
    )
    norm_gpu_indices = []
    for r in range(num_repeats):
        end_idx = loop_start + (r + 1) * period - 1
        if end_idx < len(gpu_kernels):
            norm_gpu_indices.append(end_idx)

    return NormData(
        norm_gpu_indices if len(norm_gpu_indices) >= 2 else None,
        period,
        0,
        True,
    )


def _run_eager_forward_analysis(
    eager_events: List[Dict],
    batch_type: str,
    cpu_pid: Any,
    gpu_pid: Any,
    corr_to_gpu: Dict[int, Dict],
    trace_start_ts: float,
) -> Optional[EagerForwardResult]:
    """Run eager forward analysis on companion trace for CUDAGraph attribution."""
    prefill_iters, decode_iters = collect_iteration_events(eager_events, cpu_pid)
    if batch_type == "decode":
        iters = decode_iters
        if not iters:
            iters = prefill_iters
    else:
        iters = prefill_iters
        if not iters:
            iters = decode_iters

    if not iters:
        print(f"  No {batch_type} iterations found in eager companion trace.")
        return None

    it = select_iteration(iters, eager_events, cpu_pid, 50.0, f"eager_{batch_type}")
    if it is None:
        return None

    ts = it.get("ts", 0)
    dur = it.get("dur", 0)
    print(
        f"  Eager {batch_type}: {it.get('name', '')} "
        f"(ts={ts:.0f}, from_start={format_perfetto_ts(ts, trace_start_ts)}, dur={dur:.0f}us)"
    )

    eager_gpu_pid = next(
        (e["pid"] for e in eager_events if e.get("cat") == "kernel"), None
    )
    eager_corr_to_gpu = _build_corr_to_gpu(eager_events, eager_gpu_pid)

    kernels, _, cpu_fwd = get_forward_kernels(
        eager_events, it, cpu_pid, eager_gpu_pid, corr_to_gpu=eager_corr_to_gpu
    )
    if not kernels or cpu_fwd is None:
        print(f"  No eager forward kernels found for {batch_type}.")
        return None

    mod_map = build_module_map(eager_events, cpu_fwd, kernels)
    print(f"  Eager module-mapped: {len(mod_map)}/{len(kernels)} kernels")

    norm_data = find_layer_boundaries_v2(
        eager_events, cpu_fwd, kernels, eager_gpu_pid, eager_corr_to_gpu
    )

    return EagerForwardResult(kernels, mod_map, norm_data, cpu_fwd)


def align_eager_to_graph(
    graph_kernels: List[Dict],
    eager_result: EagerForwardResult,
    label: str,
) -> Tuple[Dict[int, str], Optional[List[int]], int, int, bool]:
    """Anchor-map eager kernel metadata onto graph-mode kernel list."""
    eager_kernels = eager_result.gpu_kernels
    eager_names = [k.get("name", "") for k in eager_kernels]
    graph_names = [k.get("name", "") for k in graph_kernels]

    eager_anchor_idx: Optional[int] = None
    anchor_name: Optional[str] = None
    for ei, name in enumerate(eager_names):
        if is_stable_anchor_kernel(name):
            eager_anchor_idx = ei
            anchor_name = name
            break

    if eager_anchor_idx is None or anchor_name is None:
        print(
            f"  {label}: no stable anchor kernel found in eager trace — name-scan alignment"
        )
        eager_anchor_idx = 0
        anchor_name = eager_names[0] if eager_names else ""

    graph_anchor_idx: Optional[int] = None
    for gi, name in enumerate(graph_names):
        if name == anchor_name:
            graph_anchor_idx = gi
            break

    if graph_anchor_idx is None:
        print(f"  {label}: anchor '{anchor_name}' not found in graph — offset=0")
        offset = 0
    else:
        offset = graph_anchor_idx - eager_anchor_idx

    graph_mod_map: Dict[int, str] = {}
    for ei, mod in eager_result.module_map.items():
        gi = ei + offset
        if 0 <= gi < len(graph_kernels):
            graph_mod_map[gi] = mod
    print(
        f"  {label}: transferred {len(graph_mod_map)}/{len(graph_kernels)} module mappings"
    )

    nd = eager_result.norm_data
    if nd.norm_gpu_indices is None:
        return graph_mod_map, None, 0, nd.norm_period, nd.use_end_as_boundary

    graph_norm_indices = [
        idx + offset
        for idx in nd.norm_gpu_indices
        if 0 <= idx + offset < len(graph_kernels)
    ]
    if len(graph_norm_indices) < 4:
        print(
            f"  {label}: only {len(graph_norm_indices)} norm indices translated (need ≥4)"
        )
        return graph_mod_map, None, 0, nd.norm_period, nd.use_end_as_boundary

    pre_loop_eager = set(nd.norm_gpu_indices[: nd.norm_loop_start])
    graph_norm_loop_start = sum(
        1 for ei in pre_loop_eager if 0 <= ei + offset < len(graph_kernels)
    )
    print(
        f"  {label}: {len(graph_norm_indices)} graph norm indices, "
        f"loop_start={graph_norm_loop_start}"
    )

    return (
        graph_mod_map,
        graph_norm_indices,
        graph_norm_loop_start,
        nd.norm_period,
        nd.use_end_as_boundary,
    )


def _kernels_to_rows(
    gpu_kernels: List[Dict],
    mod_map: Dict[int, str],
    start_idx: int = 0,
) -> List[List[Any]]:
    rows = []
    for i, k in enumerate(gpu_kernels):
        cpu_mod = mod_map.get(i + start_idx, "")
        rows.append([cpu_mod, k.get("name", ""), k.get("dur", 0)])
    return rows


def slice_target_layer(
    gpu_kernels: List[Dict],
    norm_data: "NormData",
    target_layer: int,
    trace_start_ts: float,
    mod_map: Optional[Dict[int, str]] = None,
) -> List[Tuple[str, List[List[Any]]]]:
    """Slice gpu_kernels into pre-loop, target layer, and post-loop groups."""
    if mod_map is None:
        mod_map = {}
    n = len(gpu_kernels)

    nd = norm_data
    if (
        nd.norm_gpu_indices is None
        or nd.norm_period <= 0
        or len(nd.norm_gpu_indices) < 2
    ):
        print("  No layer boundaries found. Returning all kernels as a single group")
        return [("[all]", _kernels_to_rows(gpu_kernels, mod_map, 0))]

    P = nd.norm_period
    L = nd.norm_loop_start
    ngi = nd.norm_gpu_indices
    num_layers = (len(ngi) - L) // P

    if target_layer >= num_layers:
        print(
            f"  Warning: --layer {target_layer} exceeds {num_layers} layers, using the last layer"
        )
        target_layer = max(0, num_layers - 1)

    N = target_layer

    if nd.use_end_as_boundary:
        end_norm_idx = L + (N + 1) * P - 1
        layer_end = ngi[end_norm_idx] + 1 if end_norm_idx < len(ngi) else n

        if N == 0:
            layer_start = 0
        else:
            prev_end_norm_idx = L + N * P - 1
            layer_start = (
                ngi[prev_end_norm_idx] + 1 if prev_end_norm_idx < len(ngi) else 0
            )

        last_end_norm_idx = L + num_layers * P - 1
        post_start = ngi[last_end_norm_idx] + 1 if last_end_norm_idx < len(ngi) else n
    else:
        start_norm_pos = L + N * P
        end_norm_pos = L + (N + 1) * P

        layer_start = ngi[start_norm_pos] if start_norm_pos < len(ngi) else n
        layer_end = ngi[end_norm_pos] if end_norm_pos < len(ngi) else n

        last_norm_pos = L + num_layers * P
        post_start = ngi[last_norm_pos] if last_norm_pos < len(ngi) else n

    pre_end = ngi[L] if L < len(ngi) else 0

    if layer_start < n:
        first_ts = gpu_kernels[layer_start].get("ts", 0)
        print(
            f"  Layer {N}: kernels [{layer_start}:{layer_end}] of {n}, "
            f"first_ts={format_perfetto_ts(first_ts, trace_start_ts)}"
        )

    groups: List[Tuple[str, List[List[Any]]]] = []

    if pre_end > 0:
        groups.append(
            ("[pre-loop]", _kernels_to_rows(gpu_kernels[:pre_end], mod_map, 0))
        )

    groups.append(
        (
            f"[layer {N}]",
            _kernels_to_rows(gpu_kernels[layer_start:layer_end], mod_map, layer_start),
        )
    )

    if post_start < n:
        groups.append(
            (
                "[post-loop]",
                _kernels_to_rows(gpu_kernels[post_start:], mod_map, post_start),
            )
        )

    return groups


def extract_stage_kernels(
    events: List[Dict],
    iteration: Dict,
    stage_name: str,
    cpu_pid: Any,
    corr_to_gpu: Dict[int, Dict],
) -> List[List[Any]]:
    """Extract GPU kernels for a named stage (preprocess/postprocess/sample)."""
    it_ts = iteration["ts"]
    it_end = it_ts + iteration.get("dur", 0)
    it_tid = iteration.get("tid")
    it_pid = iteration.get("pid")
    target = STAGE_PREFIX + stage_name

    stage_ev = None
    for e in events:
        if e.get("ph") != "X" or e.get("cat") != "user_annotation":
            continue
        if e.get("name") != target or e.get("pid") != cpu_pid:
            continue
        if e.get("tid") != it_tid:
            continue
        e_ts = e.get("ts", 0)
        if it_ts <= e_ts and e_ts + e.get("dur", 0) <= it_end:
            stage_ev = e
            break

    if stage_ev is None:
        return []

    s_ts = stage_ev["ts"]
    s_end = s_ts + stage_ev.get("dur", 0)
    s_tid = stage_ev.get("tid")
    s_pid = stage_ev.get("pid")

    rt_launches = [
        e
        for e in events
        if e.get("ph") == "X"
        and e.get("cat") == "cuda_runtime"
        and e.get("pid") == s_pid
        and e.get("tid") == s_tid
        and s_ts <= e.get("ts", 0) <= s_end
        and is_kernel_launch(e.get("name", ""))
    ]
    rt_launches.sort(key=lambda x: x["ts"])

    matched_kernels: List[Dict] = []
    for rt in rt_launches:
        corr = (rt.get("args") or {}).get("correlation")
        if corr is not None:
            gk = corr_to_gpu.get(corr)
            if gk is not None:
                matched_kernels.append(gk)

    mod_map = build_module_map(events, stage_ev, matched_kernels)
    return _kernels_to_rows(matched_kernels, mod_map, 0)


def analyze_trace(
    events: List[Dict],
    target_layer: int,
    percentile: float,
    trace_start_ts: float,
    eager_trace_path: Optional[str],
    output_prefix: str,
) -> None:
    """Full trace analysis to XLSX."""
    cpu_pid, gpu_pid = detect_pids(events)
    if cpu_pid is None or gpu_pid is None:
        print("ERROR: Could not detect CPU/GPU pids from trace events.")
        return

    print(f"  CPU pid={cpu_pid}, GPU pid={gpu_pid}")

    corr_to_gpu = _build_corr_to_gpu(events, gpu_pid)

    prefill_iters, decode_iters = collect_iteration_events(events, cpu_pid)
    print(
        f"  Found {len(prefill_iters)} prefill and {len(decode_iters)} decode iterations"
    )

    eager_events: Optional[List[Dict]] = None
    eager_cpu_pid: Any = None
    if eager_trace_path:
        print(f"Loading eager companion trace: {eager_trace_path}")
        eager_events, _ = _load_with_cache(eager_trace_path)
        print(f"  Loaded {len(eager_events)} eager events")
        eager_cpu_pid, _ = detect_pids(eager_events)

    sections: Dict[str, List[Tuple[str, List[List[Any]]]]] = {}

    if prefill_iters:
        print()
        print("=" * 60)
        print("PREFILL ANALYSIS")
        print("=" * 60)
        prefill_iter = select_iteration(
            prefill_iters, events, cpu_pid, percentile, "prefill"
        )
        if prefill_iter:
            ts = prefill_iter["ts"]
            dur = prefill_iter.get("dur", 0)
            print(
                f"  Using: {prefill_iter.get('name', '')} "
                f"(ts={ts:.0f}, from_start={format_perfetto_ts(ts, trace_start_ts)}, dur={dur:.0f}us)"
            )
            kernels, is_cg, cpu_fwd = get_forward_kernels(
                events, prefill_iter, cpu_pid, gpu_pid, corr_to_gpu=corr_to_gpu
            )
            if kernels:
                if is_cg and eager_events is not None:
                    eager_res = _run_eager_forward_analysis(
                        eager_events,
                        "prefill",
                        eager_cpu_pid,
                        None,
                        corr_to_gpu,
                        trace_start_ts,
                    )
                    if eager_res is not None:
                        mod_map, ngi, nls, np_, use_end = align_eager_to_graph(
                            kernels, eager_res, "prefill"
                        )
                        nd = NormData(ngi, np_, nls, use_end)
                    else:
                        mod_map = {}
                        nd = _gpu_signature_fallback(kernels)
                elif not is_cg and cpu_fwd is not None:
                    mod_map = build_module_map(events, cpu_fwd, kernels)
                    print(f"  Module-mapped: {len(mod_map)}/{len(kernels)} kernels")
                    nd = find_layer_boundaries_v2(
                        events, cpu_fwd, kernels, gpu_pid, corr_to_gpu
                    )
                else:
                    mod_map = {}
                    nd = _gpu_signature_fallback(kernels)

                groups = slice_target_layer(
                    kernels, nd, target_layer, trace_start_ts, mod_map
                )
                sections["prefill"] = groups

    if decode_iters:
        print()
        print("=" * 60)
        print("DECODE ANALYSIS")
        print("=" * 60)
        decode_iter = select_iteration(
            decode_iters, events, cpu_pid, percentile, "decode"
        )
        if decode_iter:
            ts = decode_iter["ts"]
            dur = decode_iter.get("dur", 0)
            print(
                f"  Using: {decode_iter.get('name', '')} "
                f"(ts={ts:.0f}, from_start={format_perfetto_ts(ts, trace_start_ts)}, dur={dur:.0f}us)"
            )
            kernels, is_cg, cpu_fwd = get_forward_kernels(
                events, decode_iter, cpu_pid, gpu_pid, corr_to_gpu=corr_to_gpu
            )
            if kernels:
                if is_cg and eager_events is not None:
                    eager_res = _run_eager_forward_analysis(
                        eager_events,
                        "decode",
                        eager_cpu_pid,
                        None,
                        corr_to_gpu,
                        trace_start_ts,
                    )
                    if eager_res is not None:
                        mod_map, ngi, nls, np_, use_end = align_eager_to_graph(
                            kernels, eager_res, "decode"
                        )
                        nd = NormData(ngi, np_, nls, use_end)
                    else:
                        mod_map = {}
                        nd = _gpu_signature_fallback(kernels)
                elif not is_cg and cpu_fwd is not None:
                    mod_map = build_module_map(events, cpu_fwd, kernels)
                    print(f"  Module-mapped: {len(mod_map)}/{len(kernels)} kernels")
                    nd = find_layer_boundaries_v2(
                        events, cpu_fwd, kernels, gpu_pid, corr_to_gpu
                    )
                else:
                    mod_map = {}
                    nd = _gpu_signature_fallback(kernels)

                groups = slice_target_layer(
                    kernels, nd, target_layer, trace_start_ts, mod_map
                )
                sections["decode"] = groups

    stage_iter = (decode_iters[0] if decode_iters else None) or (
        prefill_iters[0] if prefill_iters else None
    )
    stage_iter = None
    if stage_iter:
        print()
        print("=" * 60)
        print("STAGE ANALYSIS")
        print("=" * 60)
        for stage_name in ("preprocess", "postprocess", "sample"):
            rows = extract_stage_kernels(
                events, stage_iter, stage_name, cpu_pid, corr_to_gpu
            )
            if rows:
                print(f"  {stage_name}: {len(rows)} kernels")
                sections[stage_name] = [(f"[{stage_name}]", rows)]
            else:
                print(f"  {stage_name}: no kernels or stage not found")

    events.clear()

    prefix = f"{output_prefix}_" if output_prefix else ""
    if sections:
        print()
        write_consolidated_xlsx(f"{prefix}trace_breakdown.xlsx", sections)
    else:
        print("No data to write.")


def _load_with_cache(filepath: str) -> Tuple[List[Dict], float]:
    """Load trace with pickle cache for fast repeated loads."""
    cache_path = Path(filepath).with_suffix("").with_suffix(".cache.pkl")
    if cache_path.exists():
        print(f"  [cache] Loading from {cache_path}")
        with cache_path.open("rb") as f:
            return pickle.load(f)

    print(f"  [cache] Parsing trace and caching to {cache_path}")
    result = load_trace_events(filepath)
    with cache_path.open("wb") as f:
        pickle.dump(result, f, protocol=pickle.HIGHEST_PROTOCOL)
    return result


def main():
    parser = argparse.ArgumentParser(
        description="Parse vLLM profiler trace to extract kernel breakdowns (v2)."
    )
    parser.add_argument("filepath", help="Path to trace JSON or JSON.GZ file")
    parser.add_argument(
        "--layer", type=int, default=8, help="Target layer/block index (default: 8)"
    )
    parser.add_argument(
        "--percentile",
        type=float,
        default=50.0,
        help="Iteration percentile to analyze 0-100 (default: 50 = median)",
    )
    parser.add_argument(
        "--eager-trace",
        default=None,
        help="Companion eager trace for CUDAGraph module attribution",
    )
    parser.add_argument(
        "--output-prefix", type=str, default="", help="Prefix for output XLSX filename"
    )
    args = parser.parse_args()

    if args.layer < 0:
        print("--layer must be >= 0")
        sys.exit(1)
    if not (0.0 <= args.percentile <= 100.0):
        print("--percentile must be 0-100")
        sys.exit(1)

    print(f"Loading trace: {args.filepath}")
    events, trace_start_ts = _load_with_cache(args.filepath)
    print(f"Loaded {len(events)} events (duration events only)")

    analyze_trace(
        events,
        target_layer=args.layer,
        percentile=args.percentile,
        trace_start_ts=trace_start_ts,
        eager_trace_path=args.eager_trace,
        output_prefix=args.output_prefix,
    )


if __name__ == "__main__":
    main()
